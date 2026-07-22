import http.client
import json
import time
import openpyxl
from openpyxl import Workbook
import folium
from folium import plugins
import numpy as np
import os  # 添加导入os模块

def get_bus_line_info(gprs_id):
    conn = http.client.HTTPSConnection("www.zjdyx.cn")
    
    payload = json.dumps({
        "gprsId": str(gprs_id),
        "dir": "true"
    })

    headers = {
        'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 MicroMessenger/7.0.20.1781(0x6700143B) NetType/WIFI MiniProgramEnv/Windows WindowsWechat/WMPF WindowsWechat(0x63090c11)XWEB/11275",
        'Content-Type': "application/json",
        'xweb_xhr': "1",
        'sec-fetch-site': "cross-site",
        'sec-fetch-mode': "cors",
        'sec-fetch-dest': "empty",
        'referer': "https://servicewechat.com/wx2c04dce60bfff2cb/33/page-frame.html",
        'accept-language': "zh-CN,zh;q=0.9"
    }

    try:
        conn.request("POST", "/WXMP_BusInfo/GetLineDetialByGprsid?sessionID=0290a7e6-8550-47bc-8331-1c6a2fc58a93", payload, headers)
        res = conn.getresponse()
        data = res.read()
        json_data = json.loads(data.decode("utf-8"))
        
        # 检查是否有效的线路数据
        if json_data.get('Item') and json_data['Item'].get('LineName'):
            return json_data
        return None
    except Exception as e:
        print(f"获取线路 {gprs_id} 数据时出错: {str(e)}")
        return None
    finally:
        conn.close()

def get_bus_shift_times(gprs_id):
    """获取线路的发车时间列表"""
    conn = http.client.HTTPSConnection("www.zjdyx.cn")
    
    payload = json.dumps({
        "gprsId": str(gprs_id),
        "dir": "false"
    })

    headers = {
        'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36 MicroMessenger/7.0.20.1781(0x6700143B) NetType/WIFI MiniProgramEnv/Windows WindowsWechat/WMPF WindowsWechat(0x63090a13) UnifiedPCWindowsWechat(0xf254160e) XWEB/18055",
        'Content-Type': "application/json",
        'xweb_xhr': "1",
        'sec-fetch-site': "cross-site",
        'sec-fetch-mode': "cors",
        'sec-fetch-dest': "empty",
        'referer': "https://servicewechat.com/wx2c04dce60bfff2cb/67/page-frame.html",
        'accept-language': "zh-CN,zh;q=0.9",
        'priority': "u=1, i"
    }

    try:
        conn.request("POST", "/WXMP_BusInfo/GetLineShiftListByGprsid?sessionID=8eea3ba5-bb1c-46a6-8953-a16159beac3d", payload, headers)
        res = conn.getresponse()
        data = res.read()
        shift_times = json.loads(data.decode("utf-8"))
        
        # 检查是否有效的发车时间数据
        if isinstance(shift_times, list) and len(shift_times) > 0:
            return shift_times
        return []
    except Exception as e:
        print(f"获取线路 {gprs_id} 发车时间时出错: {str(e)}")
        return []
    finally:
        conn.close()

def save_to_excel(json_data, workbook, route_sheet, station_sheet, original_line_num, shift_times=None):
    # 写入线路基本信息
    shift_times_str = ';'.join(shift_times) if shift_times else ''
    route_sheet.append([
        json_data['Item']['LineName'],
        json_data['Item']['LineHeading'],
        json_data['Item']['LineStartStation'],
        json_data['Item']['LineEndStation'],
        json_data['Item']['FirstShift'],
        json_data['Item']['LastShift'],
        shift_times_str,  # 发车时间列表
        original_line_num  # 添加原始线路编号
    ])
    
    # 写入站点信息
    for station in json_data['Item']['StationList']:
        station_sheet.append([
            json_data['Item']['LineName'],
            station['Index'],
            station['Name'],
            station['id'],
            station['lat'],
            station['lng'],
            original_line_num  # 添加原始线路编号
        ])

def wgs84_to_gcj02(lat, lng):
    """
    WGS84坐标系转GCJ02坐标系（火星坐标系）
    """
    a = 6378245.0  # 长半轴
    ee = 0.00669342162296594323  # 偏心率平方
    
    def transform_lat(x, y):
        ret = -100.0 + 2.0 * x + 3.0 * y + 0.2 * y * y + 0.1 * x * y + 0.2 * np.sqrt(abs(x))
        ret += (20.0 * np.sin(6.0 * x * np.pi) + 20.0 * np.sin(2.0 * x * np.pi)) * 2.0 / 3.0
        ret += (20.0 * np.sin(y * np.pi) + 40.0 * np.sin(y / 3.0 * np.pi)) * 2.0 / 3.0
        ret += (160.0 * np.sin(y / 12.0 * np.pi) + 320 * np.sin(y * np.pi / 30.0)) * 2.0 / 3.0
        return ret

    def transform_lng(x, y):
        ret = 300.0 + x + 2.0 * y + 0.1 * x * x + 0.1 * x * y + 0.1 * np.sqrt(abs(x))
        ret += (20.0 * np.sin(6.0 * x * np.pi) + 20.0 * np.sin(2.0 * x * np.pi)) * 2.0 / 3.0
        ret += (20.0 * np.sin(x * np.pi) + 40.0 * np.sin(x / 3.0 * np.pi)) * 2.0 / 3.0
        ret += (150.0 * np.sin(x / 12.0 * np.pi) + 300.0 * np.sin(x / 30.0 * np.pi)) * 2.0 / 3.0
        return ret

    if lat is None or lng is None:
        return None, None
        
    dlat = transform_lat(lng - 105.0, lat - 35.0)
    dlng = transform_lng(lng - 105.0, lat - 35.0)
    radlat = lat / 180.0 * np.pi
    magic = np.sin(radlat)
    magic = 1 - ee * magic * magic
    sqrtmagic = np.sqrt(magic)
    dlat = (dlat * 180.0) / ((a * (1 - ee)) / (magic * sqrtmagic) * np.pi)
    dlng = (dlng * 180.0) / (a / sqrtmagic * np.cos(radlat) * np.pi)
    mglat = lat + dlat
    mglng = lng + dlng
    return mglat, mglng

def create_bus_route_map(wb):
    # 创建地图，以海宁市中心为中心点
    tiles = 'https://wprd01.is.autonavi.com/appmaptile?x={x}&y={y}&z={z}&lang=zh_cn&size=1&scl=1&style=7'
    m = folium.Map(location=[30.5087, 120.6881], 
                   zoom_start=12, 
                   tiles=tiles,
                   attr='高德',  # 设置空的属性文本
                   control_scale=False,  # 不显示比例尺
                   attributionControl=False)  # 不显示属性控件
    
    # 从工作表中读取站点数据
    station_sheet = wb["站点信息"]
    
    # 用于存储所有站点的经纬度，以计算热力图
    all_locations = []
    
    # 为每条线路创建一个特征组
    line_groups = {}
    
    def get_line_color(line_num):
        try:
            num = int(line_num)
            # 根据不同范围返回不同颜色
            if 1 <= num <= 39:
                return '#FF0000'  # 鲜红色
            elif (100 <= num <= 149) or (200 <= num <= 220):  # 合并城乡公交和城乡二级公交
                return '#9400D3'  # 深紫色
            elif 150 <= num <= 199:
                return '#0000CD'  # 深蓝色
            elif 301 <= num <= 339:
                return '#006400'  # 深绿色
            elif num >= 10000:
                return '#FF1493'  # 深粉色
            else:
                return '#000000'  # 黑色
        except:
            return '#000000'  # 无法解析数字时使用黑色
    
    # 跳过表头
    for row in list(station_sheet.rows)[1:]:
        line_name = row[0].value
        original_line_num = row[6].value  # 读取原始线路编号
        lat = float(row[4].value)
        lng = float(row[5].value)
        station_name = row[2].value
        
        # 转换为高德坐标
        gcj_lat, gcj_lng = wgs84_to_gcj02(lat, lng)
        
        if gcj_lat is not None and gcj_lng is not None:
            # 使用转换后的坐标
            all_locations.append([gcj_lat, gcj_lng])
            
            # 为每条线路创建或获取特征组
            if line_name not in line_groups:
                line_groups[line_name] = folium.FeatureGroup(name=line_name)
                line_groups[line_name].locations = []  # 初始化线路经纬度列表
            
            # 添加站点标记
            folium.CircleMarker(
                location=[gcj_lat, gcj_lng],
                radius=4,
                popup=f"{line_name}: {station_name}",
                color='black',
                fill=True
            ).add_to(line_groups[line_name])
            
            # 将站点经纬度添加到线路路径
            line_groups[line_name].locations.append((gcj_lat, gcj_lng))

    # 绘制线路路径
    for line_name, group in line_groups.items():
        if group.locations:  # 确保有经纬度数据
            # 从station_sheet中获取该线路的原始编号
            for row in list(station_sheet.rows)[1:]:
                if row[0].value == line_name:
                    original_line_num = row[6].value
                    break
            folium.PolyLine(
                locations=group.locations, 
                color=get_line_color(original_line_num), 
                weight=2
            ).add_to(group)

    # 添加热力图层
    if all_locations:
        plugins.HeatMap(all_locations).add_to(m)
    
    # 添加所有线路图层
    for group in line_groups.values():
        group.add_to(m)
    
    # 在添加图层控制之前，添加图例
    legend_html = '''
    <div id="legend" style="position: fixed; 
                bottom: 50px; 
                right: 50px; 
                z-index: 1000; 
                background-color: white;
                padding: 10px;
                border: 2px solid grey;
                border-radius: 5px;
                font-size: 14px;
                max-width: 200px;">
        <div style="margin-bottom: 5px; font-weight: bold;">图例 (点击切换显示)</div>
        <div class="legend-item" data-category="city" style="cursor: pointer; margin: 2px 0;">
            <span class="legend-color" style="color: #FF0000;">■</span> <span class="legend-text">城市公交</span>
        </div>
        <div class="legend-item" data-category="rural" style="cursor: pointer; margin: 2px 0;">
            <span class="legend-color" style="color: #9400D3;">■</span> <span class="legend-text">城乡公交</span>
        </div>
        <div class="legend-item" data-category="inter" style="cursor: pointer; margin: 2px 0;">
            <span class="legend-color" style="color: #0000CD;">■</span> <span class="legend-text">县际公交</span>
        </div>
        <div class="legend-item" data-category="town" style="cursor: pointer; margin: 2px 0;">
            <span class="legend-color" style="color: #006400;">■</span> <span class="legend-text">乡镇公交</span>
        </div>
        <div class="legend-item" data-category="community" style="cursor: pointer; margin: 2px 0;">
            <span class="legend-color" style="color: #FF1493;">■</span> <span class="legend-text">社区巴士</span>
        </div>
        <div class="legend-item" data-category="other" style="cursor: pointer; margin: 2px 0;">
            <span class="legend-color" style="color: #000000;">■</span> <span class="legend-text">其他线路</span>
        </div>
    </div>
    
    <script>
        // 为图例项添加点击事件
        document.addEventListener('DOMContentLoaded', function() {
            const legendItems = document.querySelectorAll('.legend-item');
            const categoryMap = {
                'city': { ranges: [[1, 39]], color: '#FF0000' },
                'rural': { ranges: [[100, 149], [200, 220]], color: '#9400D3' },
                'inter': { ranges: [[150, 199]], color: '#0000CD' },
                'town': { ranges: [[301, 339]], color: '#006400' },
                'community': { ranges: [[10000, 10100]], color: '#FF1493' },
                'other': { ranges: [], color: '#000000' }  // 其他线路单独处理
            };
            
            legendItems.forEach(item => {
                item.addEventListener('click', function() {
                    const category = this.dataset.category;
                    const isActive = !this.classList.contains('inactive');
                    
                    // 切换激活状态
                    if (isActive) {
                        this.classList.add('inactive');
                        this.style.opacity = '0.5';
                    } else {
                        this.classList.remove('inactive');
                        this.style.opacity = '1';
                    }
                    
                    // 显示或隐藏对应类别的线路
                    toggleCategoryLines(category, !isActive);
                });
            });
        });
        
        function toggleCategoryLines(category, show) {
            const categoryMap = {
                'city': { ranges: [[1, 39]], color: '#FF0000' },
                'rural': { ranges: [[100, 149], [200, 220]], color: '#9400D3' },
                'inter': { ranges: [[150, 199]], color: '#0000CD' },
                'town': { ranges: [[301, 339]], color: '#006400' },
                'community': { ranges: [[10000, 10100]], color: '#FF1493' },
                'other': { ranges: [], color: '#000000' }
            };
            
            const config = categoryMap[category];
            if (!config) return;
            
            // 查找所有地图上的折线和圆点元素
            const paths = document.querySelectorAll('path[stroke]');
            const circles = document.querySelectorAll('circle');
            
            // 对于非"其他"类别，直接根据颜色匹配
            if (category !== 'other') {
                paths.forEach(path => {
                    if (path.getAttribute('stroke') === config.color) {
                        path.style.display = show ? 'block' : 'none';
                    }
                });
                
                circles.forEach(circle => {
                    if (circle.getAttribute('fill') === config.color) {
                        circle.style.display = show ? 'block' : 'none';
                    }
                });
            } else {
                // 对于"其他"类别，隐藏所有不匹配其他类别的黑色元素
                const otherColors = ['#FF0000', '#9400D3', '#0000CD', '#006400', '#FF1493'];
                
                paths.forEach(path => {
                    const strokeColor = path.getAttribute('stroke');
                    if (strokeColor === '#000000' && !otherColors.includes(strokeColor)) {
                        path.style.display = show ? 'block' : 'none';
                    }
                });
                
                circles.forEach(circle => {
                    const fillColor = circle.getAttribute('fill');
                    if (fillColor === '#000000' && !otherColors.includes(fillColor)) {
                        circle.style.display = show ? 'block' : 'none';
                    }
                });
            }
        }
    </script>
    '''
    m.get_root().html.add_child(folium.Element(legend_html))
    
    # 添加图层控制
    folium.LayerControl().add_to(m)
    
    # 保存地图
    m.save('bus_routes_map.html')
    print('地图已保存到 bus_routes_map.html 文件中')

# 定义线路编号范围
LINE_RANGES = [
    range(1, 40),        # 1-39
    range(100, 221),     # 100-220
    range(301, 360),     # 301-359
    [388],              # 388
    range(529, 532),    # 529-531
    [802],              # 游2
    [868],               # 868
    range(10000, 10100)   # 10000-10099
]

# 检查文件是否存在
if os.path.exists('all_bus_data.xlsx'):
    wb = openpyxl.load_workbook('all_bus_data.xlsx')  # 读取已存在的Excel文件
    print('已加载现有的 all_bus_data.xlsx 文件')
else:
    # 创建Excel工作簿和工作表
    wb = Workbook()
    route_sheet = wb.active
    route_sheet.title = "线路信息"
    station_sheet = wb.create_sheet("站点信息")

    # 修改表头
    route_sheet.append(['线路名称', '方向', '始发站', '终点站', '首班车时间', '末班车时间', '发车时间列表', '原始线路编号'])
    station_sheet.append(['线路名称', '站点序号', '站点名称', '站点ID', '纬度', '经度', '原始线路编号'])

    # 轮询所有线路
    for line_range in LINE_RANGES:
        for line_num in line_range:
            print(f"正在获取线路 {line_num} 的数据...")
            
            # 尝试获取主线数据
            main_gprs_id = int(f"{line_num}1") if line_num < 10000 else line_num
            main_line_data = get_bus_line_info(main_gprs_id)
            if main_line_data:
                # 检查首末班时间是否为空
                first_shift = main_line_data['Item'].get('FirstShift') or ''
                last_shift = main_line_data['Item'].get('LastShift') or ''
                if str(first_shift).strip() and str(last_shift).strip():
                    print(f"找到主线: {main_line_data['Item']['LineName']}")
                    # 获取发车时间
                    shift_times = get_bus_shift_times(main_gprs_id)
                    save_to_excel(main_line_data, wb, route_sheet, station_sheet, line_num, shift_times)
                else:
                    print(f"跳过主线 {main_line_data['Item']['LineName']}：首末班时间为空")
            
            # 仅为非10000系列线路尝试获取支线数据
            if line_num < 10000:
                for branch_num in range(2, 7):
                    branch_gprs_id = int(f"{line_num}{branch_num}")
                    branch_line_data = get_bus_line_info(branch_gprs_id)
                    if branch_line_data:
                        # 检查首末班时间是否为空
                        first_shift = branch_line_data['Item'].get('FirstShift') or ''
                        last_shift = branch_line_data['Item'].get('LastShift') or ''
                        if str(first_shift).strip() and str(last_shift).strip():
                            print(f"找到支线: {branch_line_data['Item']['LineName']}")
                            # 获取发车时间
                            shift_times = get_bus_shift_times(branch_gprs_id)
                            save_to_excel(branch_line_data, wb, route_sheet, station_sheet, line_num, shift_times)
                        else:
                            print(f"跳过支线 {branch_line_data['Item']['LineName']}：首末班时间为空")
            
            # 添加延时避免请求过于频繁
            time.sleep(0.5)

    # 保存Excel文件
    wb.save('all_bus_data.xlsx')
    print('所有数据已保存到 all_bus_data.xlsx 文件中')

# 绘制地图
create_bus_route_map(wb)